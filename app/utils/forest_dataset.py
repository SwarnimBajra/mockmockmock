from __future__ import annotations

from collections import Counter
from io import BytesIO
from math import log2
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DATASET_PATH = PROJECT_ROOT / "data" / "birds_dataset.xlsx"
LEGACY_DATASET_PATH = PROJECT_ROOT / "0009156-260519110011954.xlsx"
STATUS_VALUES = {"PRESENT", "ABSENT"}

if not DEFAULT_DATASET_PATH.exists() and LEGACY_DATASET_PATH.exists():
    DEFAULT_DATASET_PATH = LEGACY_DATASET_PATH


def _clean_text(value) -> str:
    return " ".join(str(value).replace("\t", " ").split()).strip()


def _parse_individual_count(value) -> int:
    text = _clean_text(value)

    if not text:
        return 1

    try:
        return max(int(float(text)), 1)
    except ValueError:
        return 1


def _find_country_index(parts: list[str]) -> int:
    for index in range(12, len(parts)):
        value = parts[index].strip()
        if len(value) == 2 and value.isalpha() and value.isupper():
            return index

    raise ValueError("country code not found")


def _find_status_index(parts: list[str], country_index: int) -> int:
    for index in range(len(parts) - 1, country_index, -1):
        if parts[index].strip() in STATUS_VALUES:
            return index

    raise ValueError("occurrence status not found")


def _parse_row(parts: list[str]) -> dict | None:
    if len(parts) < 20:
        return None

    try:
        country_index = _find_country_index(parts)
        status_index = _find_status_index(parts, country_index)
    except ValueError:
        return None

    species = _clean_text(parts[9]) if len(parts) > 9 else ""
    if not species:
        return None

    locality = _clean_text(" ".join(parts[country_index + 1 : status_index - 1]))
    state_province = _clean_text(parts[status_index - 1])
    country_code = _clean_text(parts[country_index])
    individual_count = _parse_individual_count(parts[status_index + 1] if status_index + 1 < len(parts) else "")

    return {
        "species": species,
        "country_code": country_code,
        "locality": locality,
        "state_province": state_province,
        "individual_count": individual_count,
    }


def load_occurrences(source: str | Path | bytes) -> list[dict]:
    if isinstance(source, (str, Path)):
        workbook = load_workbook(source, data_only=True, read_only=True)
    else:
        workbook = load_workbook(BytesIO(source), data_only=True, read_only=True)

    worksheet = workbook[workbook.sheetnames[0]]
    occurrences: list[dict] = []
    header_seen = False

    for row in worksheet.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue

        parts = "\t".join(str(value) for value in row if value is not None).split("\t")
        if not header_seen:
            header_seen = True
            continue

        occurrence = _parse_row(parts)
        if occurrence is not None:
            occurrences.append(occurrence)

    workbook.close()
    return occurrences


def filter_occurrences(occurrences: list[dict], area: str | None, area_field: str) -> list[dict]:
    if not area:
        return occurrences

    needle = area.casefold().strip()

    if area_field == "any":
        return [
            occurrence
            for occurrence in occurrences
            if any(
                needle in occurrence[field].casefold()
                for field in ("locality", "state_province", "country_code")
                if occurrence[field]
            )
        ]

    field_name = {
        "locality": "locality",
        "stateProvince": "state_province",
        "countryCode": "country_code",
    }[area_field]

    return [occurrence for occurrence in occurrences if needle in occurrence[field_name].casefold()]


def summarize_occurrences(occurrences: list[dict]) -> dict:
    if not occurrences:
        return {
            "species_richness": 0,
            "estimated_bird_count": 0,
            "shannon_index": 0,
            "dominance_score": 0,
            "forest_health_score": 0,
            "forest_health_label": "insufficient data",
            "species_breakdown": [],
        }

    species_counts = Counter(occurrence["species"] for occurrence in occurrences)
    bird_counts = Counter()
    for occurrence in occurrences:
        bird_counts[occurrence["species"]] += occurrence["individual_count"]

    total_birds = sum(bird_counts.values())
    unique_species = len(species_counts)

    shannon_index = 0.0
    for count in bird_counts.values():
        proportion = count / total_birds
        shannon_index -= proportion * log2(proportion)

    dominant_species, dominant_count = bird_counts.most_common(1)[0]
    dominance_score = dominant_count / total_birds

    richness_score = min(unique_species / 50, 1.0)
    diversity_score = min(shannon_index / 4, 1.0)
    balance_score = 1 - dominance_score
    forest_health_score = round((0.4 * richness_score + 0.35 * diversity_score + 0.25 * balance_score) * 100, 1)

    if forest_health_score >= 70:
        forest_health_label = "healthy"
    elif forest_health_score >= 40:
        forest_health_label = "moderate"
    else:
        forest_health_label = "degraded"

    species_breakdown = [
        {
            "species": species,
            "observations": species_counts[species],
            "estimated_birds": bird_counts[species],
        }
        for species, _ in bird_counts.most_common(10)
    ]

    return {
        "species_richness": unique_species,
        "estimated_bird_count": total_birds,
        "shannon_index": round(shannon_index, 3),
        "dominance_score": round(dominance_score, 3),
        "dominant_species": dominant_species,
        "forest_health_score": forest_health_score,
        "forest_health_label": forest_health_label,
        "species_breakdown": species_breakdown,
    }


def analyze_dataset(source: str | Path | bytes, area: str | None = None, area_field: str = "any") -> dict:
    occurrences = load_occurrences(source)
    filtered_occurrences = filter_occurrences(occurrences, area, area_field)
    summary = summarize_occurrences(filtered_occurrences)

    return {
        "records_read": len(occurrences),
        "records_used": len(filtered_occurrences),
        "summary": summary,
    }